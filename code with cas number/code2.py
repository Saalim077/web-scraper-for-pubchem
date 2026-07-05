import time
import pandas as pd
import requests
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook

# Constants
INPUT_EXCEL = "Book1.xlsx"
OUTPUT_FILE = "pubchem_results_by_cas_6.xlsx"
RETRIES = 5
NUM_THREADS = os.cpu_count() * 2
BATCH_SIZE = 100
HEADERS = {"User-Agent": "Mozilla/5.0"}

# Prepare CAS list
if not os.path.exists(INPUT_EXCEL):
    raise FileNotFoundError(f"Input file '{INPUT_EXCEL}' not found.")

df_input = pd.read_excel(INPUT_EXCEL)
cas_numbers = df_input.iloc[:, 0].dropna().astype(str).tolist()

# Create output file if not exists
if not os.path.exists(OUTPUT_FILE):
    pd.DataFrame(columns=[
        "Input CAS", "Name", "CID", "Formula", "Weight", "IUPAC", "CAS", "Synonyms"
    ]).to_excel(OUTPUT_FILE, index=False)

def get_pubchem_data_by_cas(cas_number):
    prop_url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{cas_number}/property/MolecularFormula,MolecularWeight,IUPACName/JSON"
    syn_url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{cas_number}/synonyms/JSON"

    for attempt in range(RETRIES):
        try:
            # Get properties
            prop_response = requests.get(prop_url, headers=HEADERS, timeout=10)
            if prop_response.status_code == 404:
                return None
            prop_response.raise_for_status()
            props = prop_response.json()["PropertyTable"]["Properties"][0]

            # Get synonyms
            try:
                syn_response = requests.get(syn_url, headers=HEADERS, timeout=10)
                syn_response.raise_for_status()
                synonyms = syn_response.json()["InformationList"]["Information"][0].get("Synonym", [])
            except:
                synonyms = []

            extracted_cas = next((s for s in synonyms if re.fullmatch(r"\d{2,7}-\d{2}-\d", s)), "Not found")

            return {
                "Input CAS": cas_number,
                "Name": props.get("IUPACName", "Unknown"),
                "CID": props.get("CID"),
                "Formula": props.get("MolecularFormula"),
                "Weight": props.get("MolecularWeight"),
                "IUPAC": props.get("IUPACName"),
                "CAS": extracted_cas,
                "Synonyms": ", ".join(f'"{s}"' for s in synonyms[:100])
            }

        except Exception as e:
            print(f"Attempt {attempt + 1} failed for {cas_number}: {e}")
            time.sleep(2 ** attempt)

    print(f"All retries failed for CAS {cas_number}")
    return None

def write_to_excel(data_rows):
    df = pd.DataFrame(data_rows)

    if os.path.exists(OUTPUT_FILE):
        book = load_workbook(OUTPUT_FILE)
        sheet = book.active
        startrow = sheet.max_row

        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, startrow=startrow)
    else:
        df.to_excel(OUTPUT_FILE, index=False)

    print(f"Saved {len(data_rows)} records to {OUTPUT_FILE}")

def main():
    results = []
    failed = []

    with ThreadPoolExecutor(max_workers=NUM_THREADS) as executor:
        futures = {executor.submit(get_pubchem_data_by_cas, cas): cas for cas in cas_numbers}

        for future in as_completed(futures):
            cas = futures[future]
            try:
                result = future.result()
                if result:
                    results.append(result)
                    print(f"Fetched data for CAS {cas}")
                else:
                    failed.append(cas)

                if len(results) >= BATCH_SIZE:
                    write_to_excel(results)
                    results.clear()

            except Exception as e:
                print(f"Error processing CAS {cas}: {e}")
                failed.append(cas)

    if results:
        write_to_excel(results)

    if failed:
        print(f"\nRetrying {len(failed)} failed CAS numbers...")
        retry_results = []
        with ThreadPoolExecutor(max_workers=NUM_THREADS) as executor:
            retry_futures = {executor.submit(get_pubchem_data_by_cas, cas): cas for cas in failed}
            for future in as_completed(retry_futures):
                cas = retry_futures[future]
                try:
                    result = future.result()
                    if result:
                        retry_results.append(result)
                        print(f"Retried and fetched CAS {cas}")
                except Exception as e:
                    print(f"Retry failed for CAS {cas}: {e}")

        if retry_results:
            write_to_excel(retry_results)

    print("\nAll done!")

if __name__ == "__main__":
    main()
